import pickle
import time
import PySimpleGUI as sg
from Empleado import Empleado
from Entrega import Prenda, Entrega
from datetime import date, datetime

##################################################################
##      Datos del desarrolador                                  ##
##        |-Schachner Juan Carlos                               ##
##        |-Estudienate de licenciatura en Informatica  @UNLP   ##
##        |-Cel: +54 9 2342-40-7526                             ##
##        |-Correo: juanschahcner3@gmail.com                    ##
##      Marzo 2021                                              ##
##################################################################

def menu():
    """
    Este es el menú principal del programa. Desde aquí se invocan todos los modulos 
    que forman parte del programa.Tambien se hacen pequeños controles de datos
    """
    layout = [
        [sg.T("Matadero Bragado", size=(17, 1), justification="center",
              font=("Georgia", 17))],
        [sg.T(' '),
         sg.Button('Nuevo Empleado', size=(25, 2), key="-Agregar-")],
        [sg.T(' '),
         sg.Button('Lista de empleados', size=(25, 2),key="-listar-")],
        [sg.T(' '),
         sg.Button('Entregar ropa', size=(25, 2), key="-EntregarRopa-")],
        [sg.T(' '),
         sg.Button('Ver entregas de ropa', size=(25, 2), key="-listarRopa-")],
        [sg.T(' '),
         sg.Button('Actualizar datos empleado', size=(25, 2), key="-modifEmp-")],
        [sg.T(' '),
         sg.Button('Actualizar datos entrega', size=(25, 2), key="-ModifEntrega-")],
        [sg.T(' '), sg.T("  ", size=(17, 1), justification="center")]
        ]

    window = sg.Window("Menu Principal").Layout(layout)

    while True:
        event, _values = window.Read()
        if event is None:
            break
        elif event == "-Agregar-":
            try:
                empleados= cargarDatosEmpleados() #Cargo la lista de empleados
                nuevoEmpleado=AgregarEmpleado() #Se cargan los datos del nuevo empleado
                if nuevoEmpleado == None:
                    sg.popup('No se guardó ningun empleado', title= 'Notificación')
                else:
                    agregar = True
                    for empleado in empleados:
                        if empleado.getDni() == nuevoEmpleado.getDni():
                            agregar = False
                            if empleado.getEstado():
                                sg.popup('El empleado ya se encuentra cargado no es necesario efectuar cambios', title= 'Notificación')
                            else:
                                empleado.setEstado(True)
                                guardarDatosEmpleados(empleados)
                                sg.popup('El empleado ya estaba cargado \n Pero estaba inactivo, ahora está activo', title= 'Notificación')
                    if agregar:
                        empleados.append(nuevoEmpleado) #Agrego el nuevo empleado a la lista
                        guardarDatosEmpleados(empleados) #Guardo la nueva lista de empleados
            except:
                sg.popup('Ocurrió un error inesperado \n Código 01', title= 'Error')
        elif event == "-listar-":
            try:
                empleados= cargarDatosEmpleados()
                if empleados != []:
                    listarEmpleados(empleados)
                else:
                    sg.popup('No se encontró ningun empleado en la base de datos', title= 'Notificación')
            except: 
                sg.popup('Ocurrió un error inesperado \n Código 02', title= 'Error')
        elif event == "-EntregarRopa-":
            try:
                entregarRopa()
            except:
                sg.popup('Ocurrió un error inesperado \n Código 03', title= 'Error')          
        elif event == "-listarRopa-":
            try:
                listarRopa()
            except:
                sg.popup('Ocurrió un error inesperado \n Código 04', title= 'Error')
        elif event == "-modifEmp-":
            try:
                modifEmp()
            except:
                sg.popup('Ocurrió un error inesperado \n Código 05', title= 'Error')
        elif event == '-ModifEntrega-':
            try:
                ModifEntrega()
            except:
                sg.popup('Ocurrió un error inesperado \n Código 06', title= 'Error')
    window.Close()

def ModifEntrega():
    entregas = cargarDatosEntregas()
    nombres = []
    listaPrendas = []
    miEntrega = None
    prendas = None

    for entrega in entregas:
        nombres.append(entrega.getNombre()) #genero la lista de nombres para la GUI
    layout = [
        [sg.T("Planilla de Entregas", size=(17,1), justification="center",
               font=("Georgia", 17))],
        [sg.T("Elija el empleado ", font=("Georgia", 12))],
        [sg.Combo(nombres, size=(25,1), default_value= 'Seleccione', enable_events=True)]        
        ]
    window = sg.Window("Modificar datos de la entrega").Layout(layout)
    sigue = False
    while True:
        event, values = window.Read()
        if event is None:
            break
        if values[0] != 'Seleccione': #Si se seleccionó un nombre
            sigue = True
            nombre = values[0] #Guardo el nombre seleccionado
            break
    window.close() #Cierro la ventana
    
    if sigue:
        i = 0
        for entrega in entregas: #Busco el empleado que se le entregó prendas 
            if entrega.getNombre() == nombre: #si encontré el empleado
                miEntrega = entrega #miEntrega es un objeto Entrega, guardo la entrega (todas)
                prendas = entrega.getEntregas()  #Obtengo una lista de las prendas entregadas 
                entregas.pop(i) #Borro mientrega de la lista de entregas 
                for prenda in prendas:
                    nombre = str(prenda.getProducto()) + ' ' +str(prenda.getFecha())
                    listaPrendas.append(nombre) #Genero una lista de prendas entregadas facilmente reconocibles
                break
            i += 1

        layout = [
            [sg.T("Planilla de Entregas", size=(17,1), justification="center",
                font=("Georgia", 17))],
            [sg.T("Elija la prenda ", font=("Georgia", 12))],
            [sg.Combo(listaPrendas, size=(25,1), default_value= 'Seleccione', enable_events=True)]        
            ]
        

        window = sg.Window("Modificar datos de la entrega").Layout(layout)
        seguir = True
        while seguir:
            event, values = window.Read()
            if event is None:
                sigue = False
                break
            if values[0] != 'Seleccione': #Si seleccionó una prenda
                i = 0
                for prenda in prendas: #por cada prenda
                    nombre = str(prenda.getProducto()) + ' ' +str(prenda.getFecha())
                    if values[0] == nombre: #Busco el objeto seleccionado
                        seleccion = prenda  #Lo guardo
                        prendas.pop(i) #Saco el objeto de la lista
                        seguir = False
                        window.close()
                        break
                    i += 1
        window.close()

    if sigue:

        vacio={0: '', 1: '', 2: '', 3:'', 4: 'Número', 5: '',
        6: 'Seleccione el empleado beneficiado'} #Genero un diccionario "vacio" para comprobar si
                                                # cargaron datos en el layout
        layout = [
            [sg.T("Entrega de ropa", size=(17,1), justification="center",
                font=("Georgia", 17))],
            [sg.T('Producto'), sg.Input(seleccion.getProducto() ,size=(17,1))],
            [sg.T('Tipo//Modelo'), sg.Input(seleccion.getTipo() , size=(17,1))],
            [sg.T('Marca'), sg.Input(seleccion.getMarca() , size=(17,1))],
            [sg.Checkbox('Marcar si tiene certificación', size=(25,1), default= seleccion.getCertificado())],
            [sg.T('cantidad'), sg.Input(seleccion.getCantidad() ,size=(17,1))],
            [sg.T('Fecha de entrega'), sg.Input(seleccion.getFecha() ,size=(17,1))],
            [sg.Button("Guardar", size=(12, 1), key="-guardar-"), sg.Button("Borrar", size=(12, 1), key="-Borrar-")]
            ]

        window = sg.Window("Modificar datos de la entrega ").Layout(layout)
        while sigue:
            event, value = window.read()
            if event is None:
                break
            elif event == "-Borrar-":
                if sg.popup_yes_no('Si borra la entrega no podrá recuperar los datos \n¿Desea eliminar la entrega para siempre?', title= 'Notificación'):
                    if len(prendas) > 0:
                        miEntrega.setEntrega(prendas)
                        entregas.append(miEntrega)
                        guardarDatosEntrega(entregas)
                    else:
                        guardarDatosEntrega(entregas)
                    break
            elif event ==  "-guardar-":
                completo= True
                for key in value:
                    if value[key] == vacio[key]: #Compurebo si todos los campos tienen datos
                        completo= False
                        break
                if completo: #Compruebo que los tipos de datos cargados 
                    if not(value[0].isnumeric()):
                        if not(value[1].isnumeric()):
                            if not(value[2].isnumeric()):
                                if value[4].isnumeric():
                                    valor = esFecha(value[5])
                                    if valor == True:
                                        dato = {} #Genero un diccionario con los datos de la prenda entregada
                                        dato['Producto'] = value[0]
                                        dato['Tipo'] = value[1]
                                        dato['Marca'] = value[2]
                                        dato['Certificado'] = value[3]
                                        dato['Cantidad'] = int(value[4])
                                        dato['Fecha de entrega'] = value[5]
                                        prenda = Prenda(dato) #Genero el objeto prenda
                                        prendas.append(prenda)
                                        miEntrega.setEntrega(prendas)
                                        entregas.append(miEntrega)
                                        guardarDatosEntrega(entregas)
                                        break
                                    else:
                                        sg.popup(valor,  title= 'Notificación') 
                                else:
                                    sg.popup('Error de datos en cantidad \n ingrese números no letras', title= 'Notificación')
                            else:
                                sg.popup('Error de datos en marca \n ingrese letras no números', title= 'Notificación') 
                        else:
                            sg.popup('Error de datos en Tipo//Modelo \n ingrese letras no números', title= 'Notificación')    
                    else:
                        sg.popup('Error de datos en producto \n ingrese letras no números', title= 'Notificación')
                else:
                    sg.popup('Complete todos los campos', title= 'Notificación')            
        window.close()

def modifEmp():
    def listaNombres():
        """
            Este módulo Genera una lista con los nombres de todos los empleados cargados
        """
        empleados = cargarDatosEmpleados() #Cargo los datos de los empleados
        empleados.sort(key=lambda e: e.getApellido()) #Ordeno por apellido
        lista = []
        for empleado in empleados:
            lista.append(empleado.getApellidoyNombre()) #Genero una lista de nombres
        return lista

    nombres = listaNombres()
    layout = [
        [sg.T("Planilla de datos", size=(17,1), justification="center",
               font=("Georgia", 17))],
        [sg.T("Elija el empleado ", font=("Georgia", 12))],
        [sg.Combo(nombres, size=(25,1), default_value= 'Seleccione', enable_events=True)]
        ]
    window = sg.Window("Modificar datos del empleado").Layout(layout)
    sigue = False
    while True:
        event, values = window.Read()
        if event is None:
            break
        if values[0] != 'Seleccione':
            sigue = True
            nombre = values[0]
            break
    window.close()

    dato={}
    dato['Estado']=True
    dato['Legajo']=0
    if sigue:
        i= 0
        listaempleado = cargarDatosEmpleados()
        for emp in listaempleado:
            if emp.getApellidoyNombre() == nombre:
                empleado = emp
                listaempleado.pop(i)
                break
            i+= 1

        nombre = empleado.getNombre()
        apellido = empleado.getApellido()
        direccion = empleado.getDireccion()
        fechaing = empleado.getFechaingreso()
        fechanac = empleado.getFechaNacimiento()
        celular = empleado.getCelular()
        dni = empleado.getDni()
        hijos = empleado.getHijos()
        vac= empleado.getVacunado()
        act = empleado.getEstado()
        layout = [
            [sg.T("Planilla de datos", size=(17,1), justification="center",
                font=("Georgia", 17))],
            [sg.T('Nombre'), sg.Input(nombre,size=(17,1))],
            [sg.T('Apellido'), sg.Input(apellido,size=(17,1))],
            [sg.T('Direccion'), sg.Input(direccion,size=(17,1))],
            [sg.T('Celular'), sg.Input(celular,size=(17,1))],
            [sg.T('D.N.I'), sg.Input(dni,size=(17,1))],
            [sg.T('Fecha de ingreso'), sg.Input(fechaing, size=(17,1))],
            [sg.T('Fecha de nacimiento'), sg.Input(fechanac,size=(17,1))],
            [sg.T('Cantidad de hijos'), sg.Input(hijos,size=(17,1))],
            [sg.Checkbox('Marcar si tiene todas las vacunas', vac,size=(25,1))],
            [sg.Checkbox('Marcar si el empleado esta activo', act,size=(25,1))],
            [sg.Button("Guardar", size=(12, 1), key="-guardar-"),sg.Button("BORRAR", size=(12, 1), key="-Borrar-")]
        ]
        window = sg.Window("Modificar datos del empleado").Layout(layout)
    while sigue:
        event, values = window.Read()
        
        if event is None:
            window.close()
            sg.popup('No se guardó ningun cambio', title= 'Notificación')
            break
        elif event == '-Borrar-':
            if sg.popup_yes_no('Si borra el empleado no podrá recuperar los datos \n¿Desea eliminar el empleado para siempre? \nRecuerde que puede dar de baja en vez de eliminar', title= 'Notificación'):
                guardarDatosEmpleados(listaempleado)
                break
        elif event == "-guardar-" :
            vacio={}
            vacio[0]= ""
            vacio[1]= ""
            vacio[2]= ""
            vacio[3]= ""
            vacio[4]= '(sin puntos)'
            vacio[5]= 'DD/MM/AAAA'
            vacio[6]= 'DD/MM/AAAA'
            vacio[7]= 'Número'
            vacio[8]= 'Si'
            vacio[9]= 'indiferente'
            incompleto=False
            for key in values:
                if values[key] == vacio[key]:
                    incompleto=True
                    break 
            if not incompleto :
                if not(values[0].isnumeric()):
                    if not(values[1].isnumeric()):
                        if not(values[2].isnumeric()):
                            if values[3].isnumeric():
                                if values[4].isnumeric():
                                    valor = esFecha(values[5])
                                    if valor == True:
                                        valor = esFecha(values[6])
                                        if valor == True:
                                            if values[7].isnumeric():
                                                dato['Nombre'] = values[0]
                                                dato['Apellido'] = values[1]
                                                dato['Direccion']=values[2]
                                                dato['Celular']= int(values[3])
                                                dato['D.N.I']=int(values[4])
                                                dato['Fecha de ingreso']=values[5]
                                                dato['Fecha de nacimiento']=values[6]
                                                dato['Hijos']= int(values[7])
                                                dato['Vacuna']= values[8]
                                                dato['Estado'] = values[9]
                                                nuevoEmpleado= Empleado(dato)
                                                listaempleado.append(nuevoEmpleado)
                                                guardarDatosEmpleados(listaempleado)
                                                break
                                            else:
                                                sg.popup('Tipo de dato Hijos incorrecto \n Ingrese un número', title= 'Notificación')
                                        else:
                                            sg.popup(f'Tipo de dato Fecha de nacimiento incorrecto \n {valor}', title= 'Notificación')
                                    else:
                                        sg.popup(f'Tipo de dato Fecha de ingreso incorrecto \n {valor}', title= 'Notificación')
                                else:
                                    sg.popup('Tipo de dato DNI incorrecto \n Ingrese un número', title= 'Notificación')
                            else:
                                sg.popup('Tipo de dato Celular incorrecto \n Ingrese un número', title= 'Notificación')
                        else:
                            sg.popup('Tipo de dato Dirección incorrecto \n Ingrese una Dirección', title= 'Notificación')
                    else:
                        sg.popup('Tipo de dato Apellido incorrecto \n Ingrese un apellido', title= 'Notificación')
                else:
                    sg.popup('Tipo de dato Nombre incorrecto \n Ingrese un nombre', title= 'Notificación')
            else:
                sg.popup('Complete todos los campos de datos', title= 'Notificación')
    window.close()

def listarRopa():
    def imprimir(listaEntregas,window):
        string= ''
        for entrega in listaEntregas:
            dato = entrega.getDatos()
            nombre = dato['Nombre'] 
            dni = dato['DNI']
            string+= f'El empleado {nombre}, DNI: {dni} recibió: \n'
            for prenda in dato['Entregas']:
                aux= prenda.datos()
                prod = aux['Producto']
                marc = aux['Marca']
                if aux['Certificado']:
                    cer= 'si'
                else:
                    cer= 'no'
                cant= aux['Cantidad']
                fecha = aux['Fecha de entrega']
                string += f'    {cant} {prod} marca {marc} que {cer} está certificado, el {fecha} \n'
            string += '\n-------------------------------------------------------------------------------------------------------\n'
        window['-OUTPUT-'].update(string)
            
    def ordenApellido(entregas, descendente=False):
        entregas.sort(key=lambda e: e.getNombre(), reverse=descendente)
        return entregas

    def ordenDNI(entregas, descendente=False):
        entregas.sort(key=lambda e: e.getDni(), reverse=descendente)
        return entregas

    def ordenFecha(entregas, descendente=False):
        entregas.sort(key=lambda e: e.getEntregas()[-1].getFecha(), reverse=descendente)
        return entregas
    
    def Excel(listaEntregas):
        from openpyxl import Workbook
        from openpyxl import load_workbook
        wb = load_workbook(filename = 'Moldes\\FormularioEPP.xlsx')
        ws = wb.active
        
        for entrega in listaEntregas:
            ws["A5"] = "       Nombre y Apellido del Trabajador: " + entrega.getNombre()
            ws["K5"] = "    D.N.I.: " + str(entrega.getDni())
            prendasEntregadas = entrega.getEntregas()
            i=9
            for prenda in prendasEntregadas:
                if prenda.getCertificado():
                    certificado = "Si"
                else:
                    certificado = "No"
                ws[f'B{i}'] = prenda.getProducto()
                ws[f'D{i}'] = prenda.getTipo()
                ws[f'F{i}'] = prenda.getMarca()
                ws[f'H{i}'] = certificado
                ws[f'I{i}'] = prenda.getCantidad()
                ws[f'J{i}'] = prenda.getFecha()
                i += 1

            now = datetime.now()
            nombre = entrega.getNombre() + '-' + str(now.day) + '-' + str(now.month) + '-' + str(now.year) + '-' + str(now.hour) + '-' + str(now.minute) + '-' + str(now.second)
            wb.save(filename = f"FormulariosEPP empleados\\{nombre}.xlsx")
        sg.popup(f'Archivos excel generados correctamente ', title= 'Notificación' )


    layout = [[sg.T("Filtrar por",font=("Georgia", 12)),
         sg.Combo(['Apellido','DNI','Fecha'], 
         size=(15,1), default_value= 'Apellido')],
          [sg.T("Orden",font=("Georgia", 12)),
         sg.Combo(['Ascendente','Descendente'], size=(12,1), default_value= 'Ascendente'), sg.Button("Mostrar", size=(6, 1), key="-Mostrar-"),
          sg.Button("Excel", size=(6, 1), key="-Excel-")],
          [sg.Output(size=(60,30), key='-OUTPUT-')]
         ]
    window = sg.Window('Lista de entregas de indumentaria', layout)   
    entregas = cargarDatosEntregas()
    while True:
        event, value = window.read()
        if event is None:
            break
        elif (event == "-Mostrar-") :
            if value[1] == 'Descendente': 
                descendente=True
            else:
                descendente= False
            if value[0] == 'Apellido':
                imprimir(ordenApellido(entregas, descendente),window)
            elif value[0] == 'DNI':
                imprimir(ordenDNI(entregas, descendente),window)
            elif value[0] == 'Fecha':
                imprimir(ordenFecha(entregas, descendente),window)
        elif (event == '-Excel-'):
            Excel(entregas)


    window.close()

def entregarRopa():
    """
    Este módulo se encarga de realizar la lógica de la entrega de una prenda
    """
    def listaNombres():
        """
            Este módulo Genera una lista con los nombres de todos los empleados cargados
        """
        empleados = cargarDatosEmpleados() #Cargo los datos de los empleados
        empleados.sort(key=lambda e: e.getApellido()) #Ordeno por apellido
        lista = []
        for empleado in empleados:
            lista.append(empleado.getApellidoyNombre()) #Genero una lista de nombres
        return lista

    def buscarEmpleado(nombre):
        """
            Busca el empleado recibido por parámetro  y retorna un objeto
        """
        empleados = cargarDatosEmpleados() #Cargo la lista de empleados
        for empleado in empleados: #Busco la coincidencia 
            if empleado.getApellidoyNombre() == nombre:
                return empleado #retorno el dato

    def agregarPrenda(prenda, nombre):
        """
            Busca si existe una "hoja" de entrega de prenda para un empleado
            si no hay registro de entregas, genera el primer elemento Entrega
            si hay registro de entrega y el empleado tiene una entrega, se hace un 
                append de con los datos de la nueva entrega
            si hay registro de entregas y el empleado no tiene una entrega, se hace 
                un append a la lista de entregas con la nueva entrega 
        """
        empleado = buscarEmpleado(nombre)
        entregas = cargarDatosEntregas()
        entregado = False
        if entregas == []: # si no hay registro de entrega
            nuevaEntrega = Entrega(empleado.getDni(), empleado.getApellidoyNombre(), prenda) 
                #genera el primer elemento Entrega
            entregas.append(nuevaEntrega) #se lo guarda en la lista de entregas
            entregado = True
        else: #si hay registro de entrega 
            for entrega in entregas: #por cada empleado
                if entrega.getDni() == empleado.getDni(): #si el empleado tiene una entrega
                    entrega.addEntrega(prenda) #se hace un append de con los datos de la nueva entrega
                    entregado = True
                    break
        if not entregado:
            #si hay registro de entregas y el empleado no tiene una entrega
            nuevaEntrega = Entrega(empleado.getDni(), empleado.getApellidoyNombre(), prenda)
            #Genero la nueva entrega 
            entregas.append(nuevaEntrega) #hago un append con los datos
            entregado = True
        guardarDatosEntrega(entregas) #Guardo los datos
    
    today=date.today()
    hoy= str(today.day) + "/" + str(today.month) +'/' + str(today.year) #Genero la fecha
    vacio={0: '', 1: '', 2: '', 3:'', 4: 'Número', 5: '',
     6: 'Seleccione el empleado beneficiado'} #Genero un diccionario "vacio" para comprobar si
                                            # cargaron datos en el layout

    layout = [
         [sg.T("Entrega de ropa", size=(17,1), justification="center",
               font=("Georgia", 17))],
        [sg.T('Producto'), sg.Input(size=(17,1))],
        [sg.T('Tipo//Modelo'), sg.Input(size=(17,1))],
        [sg.T('Marca'), sg.Input(size=(17,1))],
        [sg.Checkbox('Marcar si tiene certificación', size=(25,1), default= True)],
        [sg.T('cantidad'), sg.Input('Número',size=(17,1))],
        [sg.T('Fecha de entrega'), sg.Input(hoy,size=(17,1))],
        [sg.Combo(listaNombres(), size=(30,1), default_value= 'Seleccione el empleado beneficiado')],
        [sg.Button("Guardar", size=(12, 1), key="-guardar-")]
        ]

    window = sg.Window("Cargar datos de la entrega ").Layout(layout)
    while True:
        event, value = window.read()
        if event is None:
            break
        else:
            completo= True
            for key in value:
                if value[key] == vacio[key]: #Compurebo si todos los campos tienen datos
                    completo= False
                    break
            if completo: #Compruebo que los tipos de datos cargados 
                if not(value[0].isnumeric()):
                    if not(value[1].isnumeric()):
                        if not(value[2].isnumeric()):
                            if value[4].isnumeric():
                                valor = esFecha(value[5])
                                if valor == True:
                                    dato = {} #Genero un diccionario con los datos de la prenda entregada
                                    dato['Producto'] = value[0]
                                    dato['Tipo'] = value[1]
                                    dato['Marca'] = value[2]
                                    dato['Certificado'] = value[3]
                                    dato['Cantidad'] = int(value[4])
                                    dato['Fecha de entrega'] = value[5]
                                    prenda = Prenda(dato) #Genero el objeto prenda
                                    agregarPrenda(prenda, value[6]) #Agrego la prenda a la lista de entregas
                                    break
                                else:
                                    sg.popup(valor,  title= 'Notificación') 
                            else:
                                sg.popup('Error de datos en cantidad \n ingrese números no letras', title= 'Notificación')
                        else:
                            sg.popup('Error de datos en marca \n ingrese letras no números', title= 'Notificación') 
                    else:
                        sg.popup('Error de datos en Tipo//Modelo \n ingrese letras no números', title= 'Notificación')    
                else:
                    sg.popup('Error de datos en producto \n ingrese letras no números', title= 'Notificación')
            else:
                sg.popup('Complete todos los campos', title= 'Notificación')            
    window.close()

def listarEmpleados(empleados):
    def ordenApellido(empleados, descendente=False):
        empleados.sort(key=lambda e: e.getApellido(), reverse=descendente)
        return empleados

    def ordenNombre(empleados, descendente=False):
        empleados.sort(key=lambda e: e.getNombre(), reverse=descendente)
        return empleados

    def ordenDNI(empleados, descendente=False):
        empleados.sort(key=lambda e: e.getDni(), reverse=descendente)
        return empleados

    def ordenHijos(empleados, descendente=False):
        empleados.sort(key=lambda e: e.getHijos(), reverse=descendente)
        return empleados
    
    def ordenCelular(empleados, descendente=False):
        empleados.sort(key=lambda e: e.getCelular(), reverse=descendente)
        return empleados
    
    def ordenIngreso(empleados, descendente=False):
        empleados.sort(key=lambda e: e.numeroFechaIngreso(), reverse=descendente)
        return empleados
    
    def ordeNacimiento(empleados, descendente=False):
        empleados.sort(key= lambda e: e.numeroFechaNacimiento(), reverse= descendente)
        return empleados

    def ordVacunado(empleados, descendente=False):
        empleados.sort(key= lambda e: e.getVacunado(), reverse= descendente)
        return empleados

    def imprimir(listaEmpleados,window):
        string=""
        for empleado in listaEmpleados:
            string+= str(empleado) + '\n--------------------------------------------\n'
        window['-OUTPUT-'].update(string)

    def generarExcel(listaEmpleados):
        from openpyxl import Workbook
        from openpyxl import load_workbook
        wb = load_workbook(filename = 'Moldes\\Reporte Empleado.xlsx')
        ws = wb.active

        i= 2
        for empleado in listaEmpleados:
            ws[f'A{i}'] = empleado.getApellido()
            ws[f'B{i}'] = empleado.getNombre()
            ws[f'C{i}'] = empleado.getDireccion()
            ws[f'D{i}'] = empleado.getCelular()
            ws[f'E{i}'] = empleado.getDni()
            ws[f'F{i}'] = empleado.getEstadoStr()
            ws[f'G{i}'] = empleado.getFechaingreso()
            ws[f'H{i}'] = empleado.getFechaNacimiento()
            ws[f'I{i}'] = empleado.getHijos()
            ws[f'J{i}'] = empleado.getLegajo()
            if empleado.getVacunado():
                ws[f'K{i}'] = "Si"
            else:
                ws[f'K{i}'] = "No"
            i+=1
        now = datetime.now()
        exacto = str(now.day) + '-' + str(now.month) + '-' + str(now.year) + '-' + str(now.hour) + '-' + str(now.minute) + '-' + str(now.second)
        wb.save(filename = f"Reporte de empleados\\{exacto}.xlsx")
        sg.popup(f'Archivo excel generado correctamente\n {exacto}' , title= 'Notificación')

    layout = [[sg.T("Filtrar por",font=("Georgia", 12)),
         sg.Combo(['Apellido','Nombre','DNI','Celular','Hijos','Vacunado','Fecha de Ingreso','Fecha de nacimiento'], 
         size=(15,1), default_value= 'Apellido')],
          [sg.T("Orden",font=("Georgia", 12)),
         sg.Combo(['Ascendente','Descendente'], size=(12,1), default_value= 'Ascendente'), sg.Button("Mostrar", size=(6, 1), key="-Mostrar-"),
          sg.Button("Excel", size=(6, 1), key="-Excel-")],
          [sg.Output(size=(60,30), key='-OUTPUT-')]
         ]
    window = sg.Window('Lista de empleados', layout)    

    while True:
        event, value = window.read()
        if event is None:
            break
        elif (event == "-Mostrar-") :
            if value[1] == 'Descendente': 
                descendente=True
            else:
                descendente= False
            if value[0] == 'Apellido':
                imprimir(ordenApellido(empleados, descendente),window)
            elif value[0] == 'Nombre':
                imprimir(ordenNombre(empleados, descendente),window)
            elif value[0] == 'DNI':
                imprimir(ordenDNI(empleados, descendente),window)
            elif value[0] == 'Celular':
               imprimir(ordenCelular(empleados, descendente),window)
            elif value[0] == 'Hijos':
                imprimir(ordenHijos(empleados, descendente),window)
            elif value[0] == 'Vacunado':
                imprimir(ordVacunado(empleados, descendente),window)
            elif value[0] == 'Fecha de Ingreso':
                imprimir(ordenIngreso(empleados, descendente),window)
            elif value[0] == 'Fecha de nacimiento':
                imprimir(ordeNacimiento(empleados, descendente),window)
        else:
            if value[1] == 'Descendente': 
                descendente=True
            else:
                descendente= False
            if value[0] == 'Apellido':
                generarExcel(ordenApellido(empleados, descendente))
            elif value[0] == 'Nombre':
                generarExcel(ordenNombre(empleados, descendente))
            elif value[0] == 'DNI':
                generarExcel(ordenDNI(empleados, descendente))
            elif value[0] == 'Celular':
               generarExcel(ordenCelular(empleados, descendente))
            elif value[0] == 'Hijos':
                generarExcel(ordenHijos(empleados, descendente))
            elif value[0] == 'Vacunado':
                generarExcel(ordVacunado(empleados, descendente))
            elif value[0] == 'Fecha de Ingreso':
                generarExcel(ordenIngreso(empleados, descendente))
            elif value[0] == 'Fecha de nacimiento':
                generarExcel(ordeNacimiento(empleados, descendente))

    window.Close()

def cargarDatosEmpleados():
    """
    Este modulo es el encargado de leer el archivo de datos de empleados
    Lee una lista de objetos Empleados 
    """
    try:
        with open('DatosEmpleados.pckl','rb')  as f:
            listaEmpleados = pickle.load(f)
            return listaEmpleados
    except:
        return []

def guardarDatosEmpleados(listaEmpleados):
    """
    Opuesto al modulo cargarDatosEmpleados()
    Este modulo es el encargado de guardar al archivo de datos de empleados
    Guardar una lista de objetos Empleados 
    """
    try:
        with open('DatosEmpleados.pckl','wb') as f:
            pickle.dump(listaEmpleados, f)  
        sg.popup('Datos guardados', title= 'Notificación')       
    except:
        sg.popup('Ocurrió un error inesperado al guardar los datos', title= 'Notificación')

def cargarDatosEntregas():
    try:
        with open('DatosEntregas.pckl','rb')  as f:
            listaEmpleados = pickle.load(f)
            return listaEmpleados
    except:
        return []

def guardarDatosEntrega(entregas):
    try:
        with open('DatosEntregas.pckl','wb') as f:
            pickle.dump(entregas, f)  
        sg.popup('Datos de la entrega actualizados', title= 'Notificación')       
    except:
        sg.popup('Ocurrió un error inesperado al guardar los datos', title= 'Notificación')

def AgregarEmpleado():
    today=date.today()
    hoy= str(today.day) + "/" + str(today.month) +'/' + str(today.year)
    dato={}
    dato['Estado']=True
    dato['Legajo']=0
    layout = [
         [sg.T("Planilla de datos", size=(17,1), justification="center",
               font=("Georgia", 17))],
        [sg.T('Nombre'), sg.Input(size=(17,1))],
        [sg.T('Apellido'), sg.Input(size=(17,1))],
        [sg.T('Direccion'), sg.Input(size=(17,1))],
        [sg.T('Celular'), sg.Input(size=(17,1))],
        [sg.T('D.N.I'), sg.Input('(sin puntos)',size=(17,1))],
        [sg.T('Fecha de ingreso'), sg.Input(hoy,size=(17,1))],
        [sg.T('Fecha de nacimiento'), sg.Input('DD/MM/AAAA',size=(17,1))],
        [sg.T('Cantidad de hijos'), sg.Input('Número',size=(17,1))],
        [sg.Checkbox('Marcar si tiene todas las vacunas', size=(25,1))],
        [sg.Button("Guardar", size=(12, 1), key="-guardar-")]
        ]

    window = sg.Window("Cargar datos del nuevo empleado").Layout(layout)

    while True:
        event, values = window.Read()
        
        if event is None:
            return None
        elif event == "-guardar-" :
            vacio={}
            vacio[0]= ""
            vacio[1]= ""
            vacio[2]= ""
            vacio[3]= ""
            vacio[4]= '(sin puntos)'
            vacio[5]= 'DD/MM/AAAA'
            vacio[6]= 'DD/MM/AAAA'
            vacio[7]= 'Número'
            vacio[8]= 'Si'
            incompleto=False
            for key in values:
                if values[key] == vacio[key]:
                    incompleto=True
                    break 
            if not incompleto :
                if not(values[0].isnumeric()):
                    if not(values[1].isnumeric()):
                        if not(values[2].isnumeric()):
                            if values[3].isnumeric():
                                if values[4].isnumeric():
                                    valor = esFecha(values[5])
                                    if valor == True:
                                        valor = esFecha(values[6])
                                        if valor == True:
                                            if values[7].isnumeric():
                                                dato['Nombre'] = values[0]
                                                dato['Apellido'] = values[1]
                                                dato['Direccion']=values[2]
                                                dato['Celular']= int(values[3])
                                                dato['D.N.I']=int(values[4])
                                                dato['Fecha de ingreso']=values[5]
                                                dato['Fecha de nacimiento']=values[6]
                                                dato['Hijos']= int(values[7])
                                                dato['Vacuna']= values[8]
                                                nuevoEmpleado= Empleado(dato)
                                                window.close()
                                                return nuevoEmpleado
                                            else:
                                                sg.popup('Tipo de dato Hijos incorrecto \n Ingrese un número', title= 'Notificación')
                                        else:
                                            sg.popup(f'Tipo de dato Fecha de nacimiento incorrecto \n {valor}', title= 'Notificación')
                                    else:
                                        sg.popup(f'Tipo de dato Fecha de ingreso incorrecto \n {valor}', title= 'Notificación')
                                else:
                                    sg.popup('Tipo de dato DNI incorrecto \n Ingrese un número', title= 'Notificación')
                            else:
                                sg.popup('Tipo de dato Celular incorrecto \n Ingrese un número', title= 'Notificación')
                        else:
                            sg.popup('Tipo de dato Dirección incorrecto \n Ingrese una Dirección', title= 'Notificación')
                    else:
                        sg.popup('Tipo de dato Apellido incorrecto \n Ingrese un apellido', title= 'Notificación')
                else:
                    sg.popup('Tipo de dato Nombre incorrecto \n Ingrese un nombre', title= 'Notificación')
            else:
                sg.popup('Complete todos los campos de datos', title= 'Notificación')

def esFecha(cadena):
    lista = cadena.split('/')
    if len(lista) == 3:
        if int(lista[0]) <= 31:
            if int(lista[1]) <= 12:
                if int(lista[2]) > 1930:
                    today=date.today() 
                    if int(lista[2]) <= today.year :
                        return True    
                    return "El año ingresado en la fecha es invalido"
                return "El año ingresado en la fecha es invalido"
            return "El mes ingresado en la fecha es invalido"
        return "El día ingresado en la fecha es invalido"
    today=date.today()
    hoy= str(today.day) + "/" + str(today.month) +'/' + str(today.year)
    return "El formato de la fecha es invalido \n use DD/MM/AAAA \n D --> día \n M --> mes \n A --> año \n Ejemplo " + hoy

if __name__ == "__main__":
    try:
        menu()
    except:
        sg.popup('Ocurrió un error inesperado \n Código 07', title= 'Error')